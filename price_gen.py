#!/usr/bin/env python3
"""
Append a 99x12 Gaussian matrix to an existing Excel sheet so that
each row's average equals the existing row "mean" already in the sheet.

- Reads the 12 row means from column B (rows 10..21). If empty, falls back to column A.
- Writes 99 columns of samples into C10..CW21 (99 columns).
- Samples ~ N(mean, (0.2*mean)^2), then shifts so the row mean equals 'mean' exactly.

Usage:
  python append_gauss_matrix.py --xlsx your_file.xlsx --sheet "Sheet1"

  python price_gen.py \
    --xlsx pidsg25-06.xlsx \
    --sheet "p1normal"


Note:
- You asked for C10..CX21, but C..CX contains 100 columns.
  For exactly 99 columns, the correct range is C..CW (C10..CW21).
"""

import argparse
import math
import random
from statistics import mean as _mean
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter, column_index_from_string

def col_letter(n: int) -> str:
    return get_column_letter(n)

def col_index(col: str) -> int:
    return column_index_from_string(col)

def build_row(mean_val: float, n_cols: int = 99, noise_frac: float = 0.20) -> list[float]:
    if mean_val is None or (isinstance(mean_val, float) and (math.isnan(mean_val))):
        return [None] * n_cols
    sigma = abs(noise_frac * float(mean_val))
    samples = [random.gauss(float(mean_val), sigma) for _ in range(n_cols)]
    # Shift so row-average == mean exactly
    row_avg = _mean(samples)
    delta = float(mean_val) - row_avg
    return [x + delta for x in samples]

def main():
    ap = argparse.ArgumentParser()
    ap.add_argument("--xlsx", required=True, help="Path to the existing .xlsx file")
    ap.add_argument("--sheet", default=None, help="Worksheet name (default: active sheet)")
    ap.add_argument("--start_row", type=int, default=10, help="Start row for writing (default: 10)")
    ap.add_argument("--n_rows", type=int, default=12, help="Number of rows to write (default: 12)")
    ap.add_argument("--start_col", default="C", help="Start column (default: C)")
    ap.add_argument("--n_cols", type=int, default=99, help="Number of columns (default: 99)")
    ap.add_argument("--seed", type=int, default=42, help="RNG seed for reproducibility")
    args = ap.parse_args()

    random.seed(args.seed)

    wb = load_workbook(args.xlsx)
    ws = wb[args.sheet] if args.sheet else wb.active

    start_r = args.start_row
    end_r   = start_r + args.n_rows - 1
    start_c_idx = col_index(args.start_col)
    end_c_idx   = start_c_idx + args.n_cols - 1
    start_c_let = col_letter(start_c_idx)
    end_c_let   = col_letter(end_c_idx)

    # Read row means from column B first; if empty, fall back to column A
    col_mean_primary = col_index("B")
    col_mean_fallback = col_index("A")

    means: list[float] = []
    for r in range(start_r, end_r + 1):
        v = ws.cell(row=r, column=col_mean_primary).value
        if v is None or (isinstance(v, str) and v.strip() == ""):
            v = ws.cell(row=r, column=col_mean_fallback).value
        if v is None or (isinstance(v, str) and v.strip() == ""):
            raise ValueError(f"Row {r}: could not locate a numeric 'mean' in column B or A.")
        try:
            means.append(float(v))
        except Exception as e:
            raise ValueError(f"Row {r}: expected numeric mean, got {v!r}") from e

    # Generate & write the matrix
    for i, r in enumerate(range(start_r, end_r + 1)):
        row_vals = build_row(means[i], n_cols=args.n_cols, noise_frac=0.20)
        for j, c_idx in enumerate(range(start_c_idx, end_c_idx + 1)):
            ws.cell(row=r, column=c_idx, value=row_vals[j])

    out_path = args.xlsx  # in-place update
    wb.save(out_path)

    print(f"Wrote {args.n_cols} cols × {args.n_rows} rows to {start_c_let}{start_r}:{end_c_let}{end_r}")
    print(f"Row means used (rows {start_r}-{end_r}): {means}")
    print(f"Saved: {out_path}")

if __name__ == "__main__":
    main()
